from decimal import Decimal

import pandas as pd
from openpyxl import load_workbook

from recon_engine.reports.report_writer import AuditLog, write_reports


def test_excel_report_file_creation(tmp_path) -> None:
    matched = pd.DataFrame(
        [
            {
                "internal_record_id": "INT1",
                "external_record_id": "BNK1",
                "transaction_reference": "UTR1",
                "amount": Decimal("100.00"),
                "currency": "INR",
                "match_type": "EXACT",
                "confidence_score": 1.0,
                "matched_at": "2026-05-22T00:00:00+00:00",
            }
        ]
    )
    exceptions = pd.DataFrame(columns=["exception_category"])
    review_queue = pd.DataFrame(
        [
            {
                "internal_record_id": "INT2",
                "external_record_id": "BNK2",
                "transaction_reference_internal": "ABC1",
                "transaction_reference_external": "ABC-1",
                "amount_internal": Decimal("100.00"),
                "amount_external": Decimal("100.50"),
                "currency_internal": "INR",
                "currency_external": "INR",
                "direction_internal": "CREDIT",
                "direction_external": "CREDIT",
                "confidence_score": 0.75,
                "reason": "similar reference",
            }
        ]
    )
    audit = AuditLog()
    audit.add("TEST", "test audit", 1)

    write_reports(tmp_path, matched, exceptions, audit, 2, 2, review_queue)

    workbook = load_workbook(tmp_path / "reconciliation_report.xlsx")
    assert workbook.sheetnames == ["Summary", "Matched", "Exceptions", "Review Queue", "Audit Log"]
